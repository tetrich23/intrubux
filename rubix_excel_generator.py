import json, sys, os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList
from openpyxl.drawing.image import Image as XLImage
import openpyxl.chart

# ── COLOURS (dark theme mapped to professional Excel palette) ──────────────
GOLD       = "C9A800"
GOLD_LIGHT = "F5E98A"
DARK_BG    = "1A1A18"
DARK2      = "2A2A26"
MID        = "3D3D38"
WHITE      = "FFFFFF"
OFF_WHITE  = "F5F3EE"
LIGHT_GREY = "E8E6E0"
MID_GREY   = "B0AEA6"
DARK_GREY  = "5A5A56"
GREEN      = "4CAF50"
GREEN_LIGHT= "E8F5E9"
RED        = "E05040"
RED_LIGHT  = "FDECEA"
ORANGE     = "E07820"
ORANGE_LT  = "FFF3E0"
BLUE       = "2563EB"
BLUE_LIGHT = "EFF6FF"
PURPLE     = "9333EA"
PURPLE_LT  = "F5F3FF"

def fill(hex_color):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

def font(bold=False, size=11, color=None, italic=False, name="Arial"):
    kwargs = dict(bold=bold, size=size, italic=italic, name=name)
    if color: kwargs["color"] = color
    return Font(**kwargs)

def border_thin(sides="all"):
    thin = Side(style="thin", color="D0CEC8")
    thick = Side(style="medium", color="C9A800")
    none = Side(style=None)
    sides_map = {
        "all":    Border(left=thin, right=thin, top=thin, bottom=thin),
        "bottom": Border(bottom=thin),
        "thick_bottom": Border(bottom=thick),
        "left_thick": Border(left=thick, bottom=thin, top=thin, right=thin),
    }
    return sides_map.get(sides, Border())

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def apply_row(ws, row, values, styles):
    for col, (val, style) in enumerate(zip(values, styles), 1):
        c = ws.cell(row=row, column=col, value=val)
        if style.get("fill"):   c.fill = fill(style["fill"])
        if style.get("font"):   c.font = style["font"]
        if style.get("align"):  c.alignment = style["align"]
        if style.get("border"): c.border = style["border"]
        if style.get("numfmt"): c.number_format = style["numfmt"]

def merge_title(ws, row, col_start, col_end, text, bg=DARK_BG, fg=GOLD, size=13):
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row, end_column=col_end)
    c = ws.cell(row=row, column=col_start, value=text)
    c.fill = fill(bg)
    c.font = font(bold=True, size=size, color=fg)
    c.alignment = align("left", "center")

def section_header(ws, row, col_start, col_end, text):
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row, end_column=col_end)
    c = ws.cell(row=row, column=col_start, value=text)
    c.fill = fill(MID)
    c.font = font(bold=True, size=10, color=GOLD_LIGHT)
    c.alignment = align("left", "center")
    c.border = border_thin("thick_bottom")

def col_header(ws, row, cols, headers, bg=DARK2, fg=GOLD_LIGHT):
    for i, h in enumerate(headers, cols):
        c = ws.cell(row=row, column=i, value=h)
        c.fill = fill(bg)
        c.font = font(bold=True, size=9, color=fg)
        c.alignment = align("center", "center", wrap=True)
        c.border = border_thin("all")

# ── LOAD DATA ─────────────────────────────────────────────────────────────
data_path = sys.argv[1] if len(sys.argv) > 1 else "dev_data.json"
if os.path.exists(data_path):
    with open(data_path) as f:
        raw = json.load(f)
    logs = raw.get("logs", [])
else:
    # demo data if no file
    import random, time
    types = ["search","search","search","chat","image","search","chat","search"]
    labels = ["łożysko SKF 6205","wiertarka Bosch 18V","klej epoksydowy","Jak szukać po zdjęciu?",
              "📷 zdjęcie","śruba M8 nierdzewna","zamienniki MRO?","filtr hydrauliczny Parker"]
    logs = []
    base = time.time() - 86400
    for i in range(40):
        t = types[i % len(types)]
        ok = random.random() > 0.12
        dur = (random.randint(8000,28000) if t=="search"
               else random.randint(600,2500) if t=="chat"
               else random.randint(2000,5000))
        tok_in  = random.randint(1200,8000) if t in ("search","image") else random.randint(300,800)
        tok_out = random.randint(400,1800) if t in ("search","image") else random.randint(100,400)
        searches = random.randint(3,6) if t=="search" else 0
        offers = random.randint(0,6) if (t=="search" and ok) else 0
        direct = random.randint(0,offers) if offers else 0
        logs.append({
            "ts": int((base + i*2000)*1000),
            "type": t,
            "label": labels[i % len(labels)],
            "status": "ok" if ok else "error",
            "durationMs": dur,
            "tokensIn": tok_in if ok else 0,
            "tokensOut": tok_out if ok else 0,
            "webSearches": searches,
            "offersFound": offers,
            "directLinks": direct,
            "fallbackLinks": offers - direct,
            "errorType": None if ok else random.choice(["Błąd parsowania JSON","Błąd HTTP API","Błąd sieci"])
        })

# ── PRE-COMPUTE STATS ──────────────────────────────────────────────────────
ok_logs  = [l for l in logs if l["status"]=="ok"]
err_logs = [l for l in logs if l["status"]=="error"]
searches = [l for l in logs if l["type"]=="search"]
chats    = [l for l in logs if l["type"]=="chat"]
images   = [l for l in logs if l["type"]=="image"]
ok_searches = [l for l in searches if l["status"]=="ok"]
hit_searches= [l for l in ok_searches if (l.get("offersFound") or 0) > 0]

def avg_ms(lst):
    return sum(l["durationMs"] for l in lst)/len(lst) if lst else 0

avg_all    = avg_ms(ok_logs)
avg_search = avg_ms([l for l in ok_logs if l["type"]=="search"])
avg_chat   = avg_ms([l for l in ok_logs if l["type"]=="chat"])
avg_image  = avg_ms([l for l in ok_logs if l["type"]=="image"])

tok_in  = sum(l.get("tokensIn",0)  for l in logs)
tok_out = sum(l.get("tokensOut",0) for l in logs)
cost    = tok_in/1e6*3.0 + tok_out/1e6*15.0

total_direct   = sum(l.get("directLinks",0)   for l in searches)
total_fallback = sum(l.get("fallbackLinks",0)  for l in searches)
total_links    = total_direct + total_fallback
link_pct = total_direct/total_links if total_links else 0

success_rate = len(hit_searches)/len(ok_searches) if ok_searches else 0
err_rate     = len(err_logs)/len(logs) if logs else 0
web_searches = sum(l.get("webSearches",0) for l in logs)

err_types = {}
for l in err_logs:
    k = l.get("errorType") or "Nieznany błąd"
    err_types[k] = err_types.get(k, 0) + 1

# ── CREATE WORKBOOK ────────────────────────────────────────────────────────
wb = Workbook()

# ══════════════════════════════════════════════════════════════════════════
# SHEET 1 — DASHBOARD
# ══════════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "📊 Dashboard"
ws.sheet_view.showGridLines = False
ws.sheet_properties.tabColor = GOLD

# Row heights
ws.row_dimensions[1].height = 8
ws.row_dimensions[2].height = 42
ws.row_dimensions[3].height = 8

for r in range(4, 60):
    ws.row_dimensions[r].height = 20

# Column widths
col_widths = [2, 22, 16, 22, 16, 22, 16, 22, 16, 2]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ── TOP BANNER ──────────────────────────────────────────────────────────
ws.merge_cells("A1:J1")
ws["A1"].fill = fill(DARK_BG)

ws.merge_cells("A2:J2")
ws["A2"].value = "RUBIX SMART SEARCH  ·  DEVELOPER ANALYTICS REPORT"
ws["A2"].fill = fill(DARK_BG)
ws["A2"].font = Font(name="Arial", bold=True, size=18, color=GOLD)
ws["A2"].alignment = align("center","center")

ws.merge_cells("A3:J3")
ws["A3"].fill = fill(GOLD)

ts_str = datetime.now().strftime("%d.%m.%Y %H:%M")
ws.merge_cells("A4:J4")
ws["A4"].value = f"Wygenerowano: {ts_str}  ·  Łącznie zdarzeń: {len(logs)}  ·  Okres: ostatnie {len(logs)} operacji"
ws["A4"].fill = fill(DARK2)
ws["A4"].font = Font(name="Arial", size=9, color=MID_GREY)
ws["A4"].alignment = align("center","center")

# ── KPI CARDS (row 6-12) ─────────────────────────────────────────────────
def kpi_card(ws, start_row, label_col, val_col, label, value, sub, accent=GOLD, bg=DARK2):
    for r in range(start_row, start_row+6):
        for c in [label_col, val_col]:
            ws.cell(r, c).fill = fill(bg)
    ws.merge_cells(start_row=start_row,   start_column=label_col, end_row=start_row,   end_column=val_col)
    ws.merge_cells(start_row=start_row+1, start_column=label_col, end_row=start_row+1, end_column=val_col)
    ws.merge_cells(start_row=start_row+2, start_column=label_col, end_row=start_row+3, end_column=val_col)
    ws.merge_cells(start_row=start_row+4, start_column=label_col, end_row=start_row+4, end_column=val_col)

    # top accent bar
    top = ws.cell(start_row, label_col)
    top.fill = fill(accent)

    lbl = ws.cell(start_row+1, label_col, label.upper())
    lbl.font = Font(name="Arial", size=8, bold=True, color=MID_GREY)
    lbl.alignment = align("center","center")

    val = ws.cell(start_row+2, label_col, value)
    val.font = Font(name="Arial", size=20, bold=True, color=WHITE)
    val.alignment = align("center","center")

    sub_c = ws.cell(start_row+4, label_col, sub)
    sub_c.font = Font(name="Arial", size=8, color=MID_GREY, italic=True)
    sub_c.alignment = align("center","center")

ws.row_dimensions[5].height = 8

kpi_card(ws, 6,  2, 3,  "Wywołania API",    len(logs),               f"🔍 {len(searches)}  💬 {len(chats)}  📷 {len(images)}", GOLD)
kpi_card(ws, 6,  4, 5,  "Śr. czas odpow.",  f"{avg_all/1000:.2f}s",  f"Search {avg_search/1000:.1f}s  ·  Chat {avg_chat/1000:.1f}s", GREEN)
kpi_card(ws, 6,  6, 7,  "Skuteczność",       f"{success_rate:.0%}",   f"{len(hit_searches)}/{len(ok_searches)} wyszukiwań z ofertami", BLUE)
kpi_card(ws, 6,  8, 9,  "Koszt API (est.)",  f"${cost:.4f}",          f"{tok_in+tok_out:,} tokenów łącznie".replace(",","\u00a0"), ORANGE)

kpi_card(ws, 13, 2, 3,  "Błędy",            len(err_logs),           f"{err_rate:.0%} współczynnik błędów", RED if err_logs else GREEN)
kpi_card(ws, 13, 4, 5,  "Tokeny wejście",   f"{tok_in:,}".replace(",","\u00a0"),  f"${tok_in/1e6*3.0:.4f} koszt", PURPLE)
kpi_card(ws, 13, 6, 7,  "Tokeny wyjście",   f"{tok_out:,}".replace(",","\u00a0"), f"${tok_out/1e6*15.0:.4f} koszt", PURPLE)
kpi_card(ws, 13, 8, 9,  "Web Searches",     web_searches,            f"~{web_searches/len(searches):.1f} / wyszukiwanie" if searches else "—", GOLD)

# ── LINK QUALITY ROW 21-26 ──────────────────────────────────────────────
ws.row_dimensions[21].height = 10
section_header(ws, 22, 2, 9, "  🔗  JAKOŚĆ LINKÓW DO OFERT")

headers = ["", "Metryka", "Wartość", "Udział %", "Status", "", "", "", ""]
for col_idx, h in enumerate(headers, 1):
    c = ws.cell(23, col_idx, h)
    c.fill = fill(MID)
    c.font = Font(name="Arial", size=9, bold=True, color=GOLD_LIGHT)
    c.alignment = align("center","center")
    if col_idx in (2,3,4,5):
        c.border = border_thin("all")

link_rows = [
    ("Bezpośredni link do oferty", total_direct, f"=C24/(C24+C25)", "✓ Dobry" if link_pct>=0.5 else "⚠ Niski", GREEN if link_pct>=0.5 else ORANGE),
    ("Fallback (Google search)",   total_fallback, f"=C25/(C24+C25)", "—" if total_fallback==0 else "ℹ Info", MID_GREY),
]
for i, (metric, val, pct_formula, status, status_color) in enumerate(link_rows, 24):
    ws.cell(i, 2, metric).font = Font(name="Arial", size=10, color=WHITE)
    ws.cell(i, 2).border = border_thin("all")
    ws.cell(i, 3, val).font = Font(name="Arial", size=10, bold=True, color=GOLD)
    ws.cell(i, 3).alignment = align("center")
    ws.cell(i, 3).border = border_thin("all")
    pct_c = ws.cell(i, 4, pct_formula)
    pct_c.number_format = "0.0%"
    pct_c.font = Font(name="Arial", size=10, color=WHITE)
    pct_c.alignment = align("center")
    pct_c.border = border_thin("all")
    stat_c = ws.cell(i, 5, status)
    stat_c.font = Font(name="Arial", size=10, bold=True, color=status_color)
    stat_c.alignment = align("center")
    stat_c.border = border_thin("all")

# Łącznie
ws.cell(26, 2, "ŁĄCZNIE").font = Font(name="Arial", size=10, bold=True, color=GOLD_LIGHT)
ws.cell(26, 2).fill = fill(MID)
ws.cell(26, 2).border = border_thin("all")
ws.cell(26, 3, "=C24+C25").font = Font(name="Arial", size=10, bold=True, color=GOLD)
ws.cell(26, 3).fill = fill(MID)
ws.cell(26, 3).alignment = align("center")
ws.cell(26, 3).border = border_thin("all")
ws.cell(26, 4, "=C24/(C24+C25)" if total_links else 0).number_format = "0.0%"
ws.cell(26, 4).font = Font(name="Arial", size=10, bold=True, color=GOLD)
ws.cell(26, 4).fill = fill(MID)
ws.cell(26, 4).alignment = align("center")
ws.cell(26, 4).border = border_thin("all")

# background fill for unused cells on rows
for r in [22,23,24,25,26]:
    for c in [1,6,7,8,9,10]:
        ws.cell(r,c).fill = fill(DARK_BG)

# ══════════════════════════════════════════════════════════════════════════
# SHEET 2 — ANALIZA CZASU ODPOWIEDZI
# ══════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("⏱ Czasy odpowiedzi")
ws2.sheet_view.showGridLines = False
ws2.sheet_properties.tabColor = "4CAF50"

for i, w in enumerate([2,30,20,20,20,20,20,2], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

ws2.row_dimensions[1].height = 8
ws2.row_dimensions[2].height = 38
ws2.merge_cells("A2:H2")
ws2["A2"].value = "⏱  ANALIZA CZASU ODPOWIEDZI  —  RUBIX SMART SEARCH"
ws2["A2"].fill = fill(DARK_BG)
ws2["A2"].font = Font(name="Arial", bold=True, size=15, color=GOLD)
ws2["A2"].alignment = align("center","center")
ws2.merge_cells("A3:H3")
ws2["A3"].fill = fill(GOLD)
ws2.row_dimensions[3].height = 4

# ── Summary table ───────────────────────────────────────────────────────
section_header(ws2, 5, 2, 7, "  PODSUMOWANIE CZASÓW WG TYPU OPERACJI")
col_header(ws2, 6, 2, ["Typ operacji", "Liczba wywołań", "Śr. czas (ms)", "Min (ms)", "Max (ms)", "Śr. czas (s)"])

type_data = [
    ("🔍 Wyszukiwanie", [l for l in ok_logs if l["type"]=="search"]),
    ("💬 Czat AI",      [l for l in ok_logs if l["type"]=="chat"]),
    ("📷 Rozpoznawanie zdjęć", [l for l in ok_logs if l["type"]=="image"]),
    ("✦ Wszystkie razem", ok_logs),
]

chart_labels = []
chart_avgs   = []

for row_i, (label, items) in enumerate(type_data, 7):
    is_total = label.startswith("✦")
    bg = MID if is_total else DARK2
    fg_col = GOLD if is_total else WHITE
    ms_vals = [l["durationMs"] for l in items]
    avg_v = sum(ms_vals)/len(ms_vals) if ms_vals else 0
    min_v = min(ms_vals) if ms_vals else 0
    max_v = max(ms_vals) if ms_vals else 0

    if not is_total:
        chart_labels.append(label)
        chart_avgs.append(round(avg_v))

    vals = [label, len(items), round(avg_v), round(min_v), round(max_v), round(avg_v/1000, 2)]
    fmts = ["", "#,##0", "#,##0", "#,##0", "#,##0", "#,##0.00"]
    for col_i, (v, fmt) in enumerate(zip(vals, fmts), 2):
        c = ws2.cell(row_i, col_i, v)
        c.fill = fill(bg)
        c.font = Font(name="Arial", size=10, bold=is_total, color=fg_col)
        c.alignment = align("center" if col_i > 2 else "left", "center")
        c.border = border_thin("all")
        if fmt: c.number_format = fmt

# ── Timeline table (last 20 ok logs) ────────────────────────────────────
section_header(ws2, 12, 2, 7, "  OSTATNIE 20 WYWOŁAŃ — TIMELINE")
col_header(ws2, 13, 2, ["Czas", "Typ", "Opis", "Czas trwania (ms)", "Status", "Web searches"])

timeline = ok_logs[:20]
for row_i, l in enumerate(timeline, 14):
    dt = datetime.fromtimestamp(l["ts"]/1000).strftime("%d.%m %H:%M:%S")
    type_label = {"search":"🔍 Search","chat":"💬 Chat","image":"📷 Image"}.get(l["type"], l["type"])
    bg = DARK2 if row_i % 2 == 0 else "232320"
    vals = [dt, type_label, str(l.get("label",""))[:40], l["durationMs"],
            "✓ OK" if l["status"]=="ok" else "✕ ERROR", l.get("webSearches",0)]
    for col_i, v in enumerate(vals, 2):
        c = ws2.cell(row_i, col_i, v)
        c.fill = fill(bg)
        c.font = Font(name="Arial", size=9, color=WHITE)
        c.alignment = align("left" if col_i in (2,4) else "center", "center")
        c.border = border_thin("all")
        if col_i == 5:
            c.number_format = "#,##0"
        if col_i == 6:
            c.font = Font(name="Arial", size=9, color=GREEN if v=="✓ OK" else RED, bold=True)

# ── BAR CHART — avg response time by type ───────────────────────────────
# Write chart data to hidden area
chart_data_start_row = 40
ws2.cell(chart_data_start_row, 2, "Typ")
ws2.cell(chart_data_start_row, 3, "Śr. czas (ms)")
for i, (lbl, avg) in enumerate(zip(chart_labels, chart_avgs), 1):
    ws2.cell(chart_data_start_row + i, 2, lbl)
    ws2.cell(chart_data_start_row + i, 3, avg)
    ws2.row_dimensions[chart_data_start_row + i].height = 16

chart = BarChart()
chart.type = "col"
chart.style = 10
chart.title = "Średni czas odpowiedzi wg typu operacji"
chart.y_axis.title = "Milisekundy (ms)"
chart.x_axis.title = "Typ operacji"
chart.shape = 4
chart.grouping = "clustered"

data_ref = Reference(ws2, min_col=3, min_row=chart_data_start_row,
                     max_row=chart_data_start_row + len(chart_labels))
cats_ref = Reference(ws2, min_col=2, min_row=chart_data_start_row+1,
                     max_row=chart_data_start_row + len(chart_labels))
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)
chart.series[0].graphicalProperties.solidFill = GOLD
chart.series[0].graphicalProperties.line.solidFill = GOLD2 = "A88900"
chart.width = 20
chart.height = 12
ws2.add_chart(chart, "B35")

# ══════════════════════════════════════════════════════════════════════════
# SHEET 3 — TOKENY & KOSZTY
# ══════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("💰 Tokeny & Koszty")
ws3.sheet_view.showGridLines = False
ws3.sheet_properties.tabColor = ORANGE

for i, w in enumerate([2,28,20,20,20,20,2], 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

ws3.merge_cells("A1:G1")
ws3.row_dimensions[1].height = 8
ws3.row_dimensions[1].height = 8
ws3.merge_cells("A2:G2")
ws3["A2"].value = "💰  TOKENY & KOSZTY API  —  RUBIX SMART SEARCH"
ws3["A2"].fill = fill(DARK_BG)
ws3["A2"].font = Font(name="Arial", bold=True, size=15, color=GOLD)
ws3["A2"].alignment = align("center","center")
ws3.row_dimensions[2].height = 38
ws3.merge_cells("A3:G3")
ws3["A3"].fill = fill(GOLD)
ws3.row_dimensions[3].height = 4

# Token summary
section_header(ws3, 5, 2, 6, "  PODSUMOWANIE ZUŻYCIA TOKENÓW")
col_header(ws3, 6, 2, ["Typ operacji", "Tokeny wejście", "Tokeny wyjście", "Łącznie", "Koszt USD"])

tok_by_type = {}
for t in ["search","chat","image"]:
    items = [l for l in logs if l["type"]==t]
    in_ = sum(l.get("tokensIn",0) for l in items)
    out_ = sum(l.get("tokensOut",0) for l in items)
    tok_by_type[t] = (in_, out_)

type_labels_tok = [("🔍 Wyszukiwanie","search"),("💬 Czat AI","chat"),("📷 Obrazy","image")]
for row_i, (label, key) in enumerate(type_labels_tok, 7):
    in_, out_ = tok_by_type.get(key, (0,0))
    cost_v = in_/1e6*3.0 + out_/1e6*15.0
    for col_i, (v, fmt) in enumerate(zip(
        [label, in_, out_, in_+out_, cost_v],
        ["","#,##0","#,##0","#,##0","$#,##0.0000"]
    ), 2):
        c = ws3.cell(row_i, col_i, v)
        c.fill = fill(DARK2)
        c.font = Font(name="Arial", size=10, color=WHITE)
        c.alignment = align("center" if col_i>2 else "left","center")
        c.border = border_thin("all")
        if fmt: c.number_format = fmt

# Total row
total_row = 10
for col_i, (v, fmt) in enumerate(zip(
    ["ŁĄCZNIE", tok_in, tok_out, tok_in+tok_out, cost],
    ["","#,##0","#,##0","#,##0","$#,##0.0000"]
), 2):
    c = ws3.cell(total_row, col_i, v)
    c.fill = fill(MID)
    c.font = Font(name="Arial", size=10, bold=True, color=GOLD)
    c.alignment = align("center" if col_i>2 else "left","center")
    c.border = border_thin("all")
    if fmt: c.number_format = fmt

# ── Pricing assumptions ─────────────────────────────────────────────────
section_header(ws3, 12, 2, 6, "  ZAŁOŻENIA CENOWE (Claude Sonnet 4.6)")
col_header(ws3, 13, 2, ["Parametr", "Wartość", "Jednostka", "Notatka"])
price_rows = [
    ("Cena tokenów wejściowych", 3.00, "USD / 1M tokenów", "Claude Sonnet 4.6 — szacunkowe"),
    ("Cena tokenów wyjściowych", 15.00, "USD / 1M tokenów", "Claude Sonnet 4.6 — szacunkowe"),
    ("Tokeny wejściowe (łącznie)", tok_in, "tokenów", "Z logów sesji"),
    ("Tokeny wyjściowe (łącznie)", tok_out, "tokenów", "Z logów sesji"),
    ("Szacowany koszt całkowity", cost, "USD", "Obliczono z tokenów × ceny"),
]
for row_i, (param, val, unit, note) in enumerate(price_rows, 14):
    is_result = row_i == 18
    bg = MID if is_result else DARK2
    c_param = ws3.cell(row_i, 2, param)
    c_param.fill = fill(bg)
    c_param.font = Font(name="Arial", size=10, color=GOLD if is_result else WHITE)
    c_param.border = border_thin("all")

    c_val = ws3.cell(row_i, 3, val)
    c_val.fill = fill(bg)
    c_val.font = Font(name="Arial", size=10, bold=True, color=GOLD if is_result else WHITE)
    c_val.alignment = align("right","center")
    c_val.number_format = "$#,##0.0000" if is_result else "#,##0.00" if isinstance(val,float) else "#,##0"
    c_val.border = border_thin("all")

    c_unit = ws3.cell(row_i, 4, unit)
    c_unit.fill = fill(bg)
    c_unit.font = Font(name="Arial", size=9, color=MID_GREY, italic=True)
    c_unit.alignment = align("center","center")
    c_unit.border = border_thin("all")

    c_note = ws3.cell(row_i, 5, note)
    c_note.fill = fill(bg)
    c_note.font = Font(name="Arial", size=9, color=DARK_GREY)
    c_note.alignment = align("left","center")
    c_note.border = border_thin("all")

# ── Pie chart — token split ─────────────────────────────────────────────
ws3.cell(35, 2, "Kategoria")
ws3.cell(35, 3, "Tokeny")
ws3.cell(36, 2, "Wejściowe"); ws3.cell(36, 3, tok_in)
ws3.cell(37, 2, "Wyjściowe"); ws3.cell(37, 3, tok_out)

pie = PieChart()
pie.title = "Podział tokenów: wejście vs wyjście"
pie.style = 10
pie_data = Reference(ws3, min_col=3, min_row=35, max_row=37)
pie_cats = Reference(ws3, min_col=2, min_row=36, max_row=37)
pie.add_data(pie_data, titles_from_data=True)
pie.set_categories(pie_cats)
pie.series[0].graphicalProperties.solidFill = GOLD
slice0 = DataPoint(idx=0); slice0.graphicalProperties.solidFill = GOLD
slice1 = DataPoint(idx=1); slice1.graphicalProperties.solidFill = ORANGE
pie.series[0].data_points = [slice0, slice1]
pie.width = 16; pie.height = 12
ws3.add_chart(pie, "B21")

# ══════════════════════════════════════════════════════════════════════════
# SHEET 4 — LOGI ZDARZEŃ
# ══════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("📋 Logi zdarzeń")
ws4.sheet_view.showGridLines = False
ws4.sheet_properties.tabColor = BLUE

col_widths4 = [2,20,14,42,18,18,14,16,16,18,2]
for i, w in enumerate(col_widths4, 1):
    ws4.column_dimensions[get_column_letter(i)].width = w

ws4.row_dimensions[1].height = 8
ws4.row_dimensions[2].height = 38
ws4.merge_cells("A2:K2")
ws4["A2"].value = "📋  PEŁNE LOGI ZDARZEŃ  —  RUBIX SMART SEARCH"
ws4["A2"].fill = fill(DARK_BG)
ws4["A2"].font = Font(name="Arial", bold=True, size=15, color=GOLD)
ws4["A2"].alignment = align("center","center")
ws4.merge_cells("A3:K3")
ws4["A3"].fill = fill(GOLD)
ws4.row_dimensions[3].height = 4

section_header(ws4, 5, 2, 10, f"  WSZYSTKIE ZDARZENIA ({len(logs)} rekordów)  ·  max. 200 w pamięci")
col_header(ws4, 6, 2, ["Data i czas","Typ","Opis","Czas trwania (ms)","Tokeny wejście","Tokeny wyjście","Web searches","Oferty","Link bezpośredni"])

for row_i, l in enumerate(logs, 7):
    ws4.row_dimensions[row_i].height = 16
    alt = row_i % 2 == 0
    base_bg = "232320" if alt else DARK2

    dt_val = datetime.fromtimestamp(l["ts"]/1000)
    type_map = {"search":"🔍 Search","chat":"💬 Chat","image":"📷 Image"}
    status_ok = l["status"] == "ok"

    type_label = type_map.get(l["type"], l["type"])
    type_color = {"search":GOLD,"chat":"818CF8","image":"EC4899"}.get(l["type"], WHITE)

    row_vals = [
        (dt_val,              "DD.MM.YYYY HH:MM:SS", WHITE,    "center"),
        (type_label,          "",                    type_color,"center"),
        (str(l.get("label",""))[:40], "",            WHITE,    "left"),
        (l.get("durationMs",0),"#,##0",              GOLD if status_ok else RED, "right"),
        (l.get("tokensIn",0), "#,##0",               WHITE,    "right"),
        (l.get("tokensOut",0),"#,##0",               WHITE,    "right"),
        (l.get("webSearches",0),"#,##0",             MID_GREY, "center"),
        (l.get("offersFound",0) if l["type"]=="search" else "—",
                              "#,##0" if l["type"]=="search" else "", WHITE, "center"),
        (l.get("directLinks",0) if l["type"]=="search" else "—",
                              "#,##0" if l["type"]=="search" else "", GREEN if l.get("directLinks",0)>0 else MID_GREY, "center"),
    ]

    row_bg = RED_LIGHT.replace("FD","2A").replace("EC","1F") if not status_ok else base_bg
    row_bg = "2A1F1E" if not status_ok else base_bg

    for col_i, (v, fmt, fg, h) in enumerate(row_vals, 2):
        c = ws4.cell(row_i, col_i, v)
        c.fill = fill(row_bg)
        c.font = Font(name="Arial", size=9, color=fg)
        c.alignment = align(h, "center")
        c.border = border_thin("all")
        if fmt: c.number_format = fmt

# ── Freeze panes & auto-filter ──────────────────────────────────────────
ws4.freeze_panes = "B7"
ws4.auto_filter.ref = f"B6:J{6+len(logs)}"

# ══════════════════════════════════════════════════════════════════════════
# SHEET 5 — BŁĘDY
# ══════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("⚠️ Błędy")
ws5.sheet_view.showGridLines = False
ws5.sheet_properties.tabColor = RED

for i, w in enumerate([2,35,18,18,35,2], 1):
    ws5.column_dimensions[get_column_letter(i)].width = w

ws5.row_dimensions[2].height = 38
ws5.merge_cells("A2:F2")
ws5["A2"].value = "⚠️  ANALIZA BŁĘDÓW  —  RUBIX SMART SEARCH"
ws5["A2"].fill = fill(DARK_BG)
ws5["A2"].font = Font(name="Arial", bold=True, size=15, color=GOLD)
ws5["A2"].alignment = align("center","center")
ws5.merge_cells("A3:F3")
ws5["A3"].fill = fill(RED)
ws5.row_dimensions[3].height = 4

# Error summary
section_header(ws5, 5, 2, 5, "  BŁĘDY WG TYPU")
col_header(ws5, 6, 2, ["Typ błędu", "Liczba", "Udział %", "Rekomendacja"])
err_recos = {
    "Błąd parsowania JSON":  "Sprawdź prompt — AI zwróciło nieprawidłowy JSON",
    "Błąd HTTP API":         "Sprawdź klucz API i limity rate",
    "Błąd sieci":            "Sprawdź połączenie internetowe",
    "Brak klucza API":       "Skonfiguruj klucz API w ustawieniach",
    "Nieznany błąd":         "Sprawdź logi w konsoli przeglądarki",
}
err_items = sorted(err_types.items(), key=lambda x: -x[1])

for row_i, (etype, cnt) in enumerate(err_items or [("Brak błędów", 0)], 7):
    pct = cnt/len(logs) if logs else 0
    reco = err_recos.get(etype, "Sprawdź logi aplikacji")
    for col_i, (v, fmt, color) in enumerate([
        (etype, "", WHITE),
        (cnt, "#,##0", RED if cnt>0 else GREEN),
        (pct, "0.0%", RED if pct>0.1 else ORANGE if pct>0 else GREEN),
        (reco, "", MID_GREY),
    ], 2):
        c = ws5.cell(row_i, col_i, v)
        c.fill = fill(DARK2)
        c.font = Font(name="Arial", size=10, color=color)
        c.alignment = align("center" if col_i in (3,4) else "left","center")
        c.border = border_thin("all")
        if fmt: c.number_format = fmt

# Error log detail
err_start = max(10, 7+len(err_items)+2)
section_header(ws5, err_start, 2, 5, f"  SZCZEGÓŁOWE LOGI BŁĘDÓW ({len(err_logs)} rekordów)")
col_header(ws5, err_start+1, 2, ["Data i czas","Typ operacji","Opis","Typ błędu"])

for row_i, l in enumerate(err_logs[:50], err_start+2):
    dt = datetime.fromtimestamp(l["ts"]/1000).strftime("%d.%m.%Y %H:%M:%S")
    type_label = {"search":"🔍 Search","chat":"💬 Chat","image":"📷 Image"}.get(l["type"],l["type"])
    for col_i, v in enumerate([dt, type_label, str(l.get("label",""))[:40], l.get("errorType","—")], 2):
        c = ws5.cell(row_i, col_i, v)
        c.fill = fill("2A1F1E")
        c.font = Font(name="Arial", size=9, color=RED if col_i==5 else WHITE)
        c.border = border_thin("all")
        c.alignment = align("center" if col_i==3 else "left","center")

if not err_logs:
    ws5.merge_cells(f"B{err_start+2}:E{err_start+2}")
    c = ws5.cell(err_start+2, 2, "🎉  Brak błędów w logach — świetna robota!")
    c.fill = fill("1A2E1A")
    c.font = Font(name="Arial", size=12, bold=True, color=GREEN)
    c.alignment = align("center","center")

# ── Bar chart for errors ─────────────────────────────────────────────────
if err_items and err_items[0][1] > 0:
    err_chart_row = err_start + 20
    ws5.cell(err_chart_row, 2, "Typ błędu")
    ws5.cell(err_chart_row, 3, "Liczba")
    for i, (etype, cnt) in enumerate(err_items, 1):
        ws5.cell(err_chart_row+i, 2, etype[:30])
        ws5.cell(err_chart_row+i, 3, cnt)

    err_chart = BarChart()
    err_chart.type = "bar"
    err_chart.style = 10
    err_chart.title = "Błędy wg typu"
    err_chart.y_axis.title = "Liczba błędów"
    err_data = Reference(ws5, min_col=3, min_row=err_chart_row, max_row=err_chart_row+len(err_items))
    err_cats = Reference(ws5, min_col=2, min_row=err_chart_row+1, max_row=err_chart_row+len(err_items))
    err_chart.add_data(err_data, titles_from_data=True)
    err_chart.set_categories(err_cats)
    err_chart.series[0].graphicalProperties.solidFill = RED
    err_chart.width = 20; err_chart.height = 10
    ws5.add_chart(err_chart, f"B{err_start+8}")

# ── Fill background cols ─────────────────────────────────────────────────
for ws_obj in [ws, ws2, ws3, ws4, ws5]:
    for r in range(1, 100):
        for c in [1, ws_obj.max_column]:
            cell = ws_obj.cell(r, c)
            if not cell.fill or cell.fill.patternType is None or cell.fill.patternType == "none":
                cell.fill = fill(DARK_BG)

# ── SAVE ──────────────────────────────────────────────────────────────────
out_path = sys.argv[2] if len(sys.argv) > 2 else "/mnt/user-data/outputs/rubix_analytics.xlsx"
wb.save(out_path)
print(f"OK:{out_path}")
